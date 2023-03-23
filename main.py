"""import time
import os
import traceback
import openpyxl
from openpyxl.utils import get_column_letter"""
from getters import *
import csv

start = time.time()

# Nom du fichier horodaté
"""dateFile = time.strftime(getFileInfo("dateFormat"))
fileType = getFileInfo("fileType")"""

path = getFileInfo("path")
filename = "%s\%s_%s.csv" % (path, getFileInfo("genre"), time.strftime(getFileInfo("dateFormat")))

# les colonnes
section = getsection()
fieldnames = getkeys(section[0])

maxRecords = int(getFileInfo("maxRecords"))

"""# création d'un classeur
wb = openpyxl.Workbook()

# création d'une feuille
ws = wb.active

# écrire les en-têtes de colonne
for i, fieldname in enumerate(fieldnames):
    col_letter = get_column_letter(i + 1)
    ws.cell(row=1, column=i+1, value=fieldname)

# écrire les données
for j in range(1, maxRecords+1):
    row = []
    for fieldname in fieldnames:
        row.append(getData(fieldname))
    ws.append(row)

# enregistrer le classeur
wb.save(filename)"""

# ouvrir le fichier csv en mode écriture
with open(filename, 'w', newline='') as csvfile:
    # créer un écrivain csv
    writer = csv.writer(csvfile)

    # écrire les en-têtes de colonne
    writer.writerow(fieldnames)

    # écrire les données
    for j in range(1, maxRecords+1):
        row = []
        for fieldname in fieldnames:
            row.append(getData(fieldname))
        writer.writerow(row)


end = time.time()
print("csv généré en " + str(end - start) + " secondes")